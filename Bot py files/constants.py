from openpyxl import load_workbook

Base_URL = "https://www.booking.com"


wb = load_workbook('Hotel_Booking_Input_Data.xlsx')
ws = wb['Data']

currency = ws['C2'].value
place_to_go = ws['C3'].value
check_in_date = ws['C4'].value
check_out_date = ws['C5'].value
adults = ws['C6'].value
